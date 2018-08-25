# -*- coding: utf-8 -*-
"""
Created on Fri May 25 11:50:07 2018

@author: zachary.shaver
"""
import os 
from pathlib import Path
import win32com.client
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime


email_list = {'Khendra': "khendra.davidson@baesystems.com",
              'Dylan': 'dylan.drum@baesystems.com', 
              'Dontavius': 'dontavius.hopkins@baesystems.com', 
              'Jaylan': 'jaylan.mobley@baesystems.com', 
              'Yuliya': 'yuliya.pinchuk@baesystems.com', 
              'Zachary': 'zachary.shaver@baesystems.com', 
              'Sean': 'sean.stewart2@baesystems.com', 
              'Alina': 'alina.svarishchuk@baesystems.com', 
              'George': 'george.vargas@baesystems.com', 
              'Darnell': 'darnell.wallace@baesystems.com',
              'Brandon': 'brandon.wells@baesystems.com'}



sch_path = 'C:/Users/Zachary.shaver/Desktop/5_30/'
sp_path = "//wtooleast.nosc.na.baesystems.com/Network_Documentation/Enterprise_Documentation/Voice/FedRAMP Migration 2018/Scripts\put_share_point_here/sharepoint.xlsx"
sp_master = {}


sp_append = ""

sp_date = input("Date :")


sp_wb = load_workbook(str(sp_path))
sp_ws = sp_wb.get_sheet_by_name(sp_wb.get_sheet_names()[0])


EMAIL = sp_ws['E']
Training_date = sp_ws['I']
Training_time = sp_ws['J']
Migration_Complete = sp_ws['K']
Schedule_metod = sp_ws['L']


for x in range(len(Training_date)):
    sp_item = {}
    sp_item.update({'Training date': Training_date[x]})
    sp_item.update({'Training time': Training_time[x]})
    sp_item.update({'Migration Complete': Migration_Complete[x]})
    sp_item.update({'Schedule method': Schedule_metod[x]})
    sp_item.update({'row': x})
    sp_master.update({EMAIL[x].value.lower(): sp_item})
    


DAY = '5_23'



def print_structure():
    for x in months.keys():
        for y in months[x].keys():
            print(x,y,list(months[x][y]))
            
def check_valid_file(fnam):
    if '-' not in fnam or ' ' in fnam or '.xlsx' not in fnam or '~' in fnam:
        return False
    else:
        return True

def pack_dt(sch_path):
#get the structure from the folder
    for hour in os.listdir(Path(sch_path)):
        #if its not a valid file it skips it 
        print(os.fsdecode(hour))
        if not check_valid_file(os.fsdecode(hour)):
            continue
        print(Path(sch_path+'/'+os.fsdecode(hour)).exists(), 'valid file')
        sch_wb = load_workbook(sch_path+'/'+os.fsdecode(hour))
        sch_ws = sch_wb.get_sheet_by_name(sch_wb.get_sheet_names()[0])
        
        #sch ws values
        mig_status = sch_ws['F']
        name = sch_ws['A']
        sch_email = sch_ws['B']
        user_number = sch_ws['C']
        
        
        time = os.fsdecode(hour).split('.')[0]
        
        #go through that schedgules emails
        for x in range(16, len(sch_email)):
            
            
            
            
            if not sch_email[x].value:
                print(x,'th item in ', os.fsdecode(hour), ' has blank email')
                continue
            
            
            print(time, os.fsdecode(hour))
            try:    
                
                #set variables in the try catch to find key errors that may come up from dict searches 
                sp_schmethod = sp_ws['L'][sp_master[sch_email[x].value.lower()]['row']].value
                sp_migvalS = sp_ws['K'][sp_master[sch_email[x].value.lower()]['row']].value
                sp_timeval = sp_ws.cell(row=sp_master[sch_email[x].value.lower()]['row']+1, column=10)
                sp_migval = sp_ws.cell(row=sp_master[sch_email[x].value.lower()]['row']+1, column=11)
                sp_traindateval = sp_ws.cell(row=sp_master[sch_email[x].value.lower()]['row']+1, column=9)
                
                #print(sch_email[x].value)
                
                if not sp_schmethod:                                                            #blank sch
                    print('empty value for learn sch_method')
                    sp_timeval.value = time #time 
                    sp_migval.value = mig_status[x].value #migration status 
                    sp_traindateval.value = sp_date #training date
                    
                elif 'manual' in sp_schmethod.lower():                                        #manual users
                    print('manual value for learn sch_method')
                    if not mig_status[x].value or 'no' in mig_status[x].value.lower(): #no for manual case (cancel out time and stuff)
                    #insert for future users
                        if not sp_migvalS:
                            print(sch_email[x].value, '(blank) in (manual) (no) in attendace case clear vals')
                            sp_timeval.value = '' #time 
                            sp_migval.value = '' #migration status 
                            sp_traindateval.value = '' #training date
                            
                        elif 'yes' in sp_migvalS.lower():
                            print(sch_email[x].value, '(yes) in (manual) but (no) in attendance pass case')
                            
                        else:
                            print(sch_email[x], '(no) in (manual) (no) in attendance case (clear vals)')
                            sp_timeval.value = '' #time 
                            sp_migval.value = '' #migration status 
                            sp_traindateval.value = '' #training date
                            
                    elif 'yes' in mig_status[x].value.lower():
                        
                        if not sp_migvalS:
                            print(sch_email[x].value, '(blank) in (manual) (yes) in attendace case enter vals')
                            sp_timeval.value = time #time 
                            sp_migval.value = mig_status[x].value #migration status 
                            sp_traindateval.value = sp_date #training date
                            
                        elif 'yes' in sp_migvalS.lower():
                            print(sch_email[x].value, '(yes) in (manual) but (yes) in attendance pass case')
                            
                        else:
                            print(sch_email[x], '(no) in (manual) (yes) in attendance case enter values')
                            sp_timeval.value = time #time 
                            sp_migval.value = mig_status[x].value #migration status 
                            sp_traindateval.value = sp_date #training date 
                            
                        
                elif 'ea' in sp_schmethod.lower() or 'fedramp' in sp_schmethod.lower():                                         #ea
                    print('ea or fedramp')
                    continue
                
                elif 'exec' in sp_schmethod.lower() or 'ilearn' in sp_schmethod.lower():                                        #exec/ilearn
                    print('exec and ilearn')    
                    if not mig_status[x].value or 'no' in mig_status[x].value.lower():
                        
                        if not sp_migvalS or 'no' in sp_migvalS.lower():
                            print(sch_email[x].value.lower(), '(no/blank) in (exec/ilearn) (yes) in attendace case enter vals')
                            sp_timeval.value = time #time 
                            sp_migval.value = mig_status[x].value #migration status 
                            sp_traindateval.value = sp_date #training date
                            
                        elif 'yes' in sp_schmethod.lower():
                            print(sch_email[x], 'anomily (no) on attendance and (yes) on sharepoint')
                            
                    elif 'yes' in mig_status[x].value.lower():
                        
                        if not sp_migvalS or 'no' in sp_migvalS.lower():
                            print(sch_email[x].value.lower(), '(no/blank) in (exec/ilearn) (yes) in attendace case enter vals')
                            sp_timeval.value = time #time 
                            sp_migval.value = mig_status[x].value #migration status 
                            sp_traindateval.value = sp_date #training date
                            
                        elif 'yes' in sp_schmethod.lower():
                            print(sch_email[x], 'anomily (no) on attendance and (yes) on sharepoint (exec/ilearn)')
                        
            except KeyError:
                print('Manual required for ' , name[x].value, sch_email[x].value, mig_status[x].value, time, os.fsdecode(hour))
            except AttributeError as e:
                print('att error dont mind this', e)


    

pack_dt(str(sch_path))
#print_structure() 
print('exited pack')
sp_wb.save(str(sp_path)[:-5] + 'test.xlsx')