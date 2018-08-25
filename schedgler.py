# -*- coding: utf-8 -*-
"""
Created on Mon Jun  4 14:35:15 2018

@author: zachary.shaver
"""
from pathlib import Path
from openpyxl import load_workbook
import os
from string import ascii_uppercase
from random import shuffle
print('FORMAT /|MONTH NAME (first letter caps)|/|MONTH NUMBER|_|DAY|/ EX: /June/6_10/')
date_to_sch = input('Enter date: ')
print('ENTER DATE IN THIS FORMAT : Thursday, July 5, 2018')
current_date = input('Enter date: ')
print('GRABBING TRAINERS SCHEDGULE.....')

month_folder = '//wtooleast.nosc.na.baesystems.com/Network_Documentation/Enterprise_Documentation\Voice\FedRAMP Migration 2018/Sign Up  Lists'
sp_path = '//wtooleast.nosc.na.baesystems.com/Network_Documentation/Enterprise_Documentation/Voice\FedRAMP Migration 2018\Scripts\put_share_point_here\sharepoint.xlsx'
time_path = "//wtooleast.nosc.na.baesystems.com/Network_Documentation/Enterprise_Documentation/Voice/FedRAMP Migration 2018/Scripts/times.xlsx"

sp_master = {}
manual_users = {}
sp_wb = load_workbook(str(sp_path))
sp_ws = sp_wb.get_sheet_by_name(sp_wb.get_sheet_names()[0])


EMAIL = sp_ws['E']
num_meeting = sp_ws['G']
trainers = {'Khendra' : {'t' : 0, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'Dylan' : {'t' : 0, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'Dontavius' : {'t' : 1, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'Yuliya' : {'t' : 1, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'Zachary': {'t' : 2, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'Sean' : {'t' : 2, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'Alina' : {'t' : 1, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'George' : {'t' : 1, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'Darnell' : {'t' : 2, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0},
            'Brandon' : {'t' : 1, 'start time' : '8', 'num_meetings': 0, 'total_res' : 0}}

trainer_list = ['Khendra','Dylan', 'Dontavius', 'Yuliya', 'Zachary', 'Sean', 'Alina', 'George', 'Darnell','Brandon']



def pack_times(day):
    time_wb = load_workbook(time_path)
    time_ws = time_wb.get_sheet_by_name(time_wb.get_sheet_names()[0])
    time_trainers = time_ws['A']
    time_day = None
    all_days = []
    for x in time_ws.iter_cols():
        #print(x[0].value)
        if not x[0].value:
            continue
        all_days.append(x[0].value)
        if day in x[0].value:
            time_day = x
            break
    try:
        
        for x in range(1, len(time_trainers)):
            if 't' in time_day[x].value:
                trainers[time_trainers[x].value]['t'] = 1
                trainers[time_trainers[x].value]['start time'] = time_day[x].value[1]
            elif 's' in time_day[x].value or 'l' in time_day[x].value or 'p' in time_day[x].value:
                trainers[time_trainers[x].value]['t'] = 2
                trainers[time_trainers[x].value]['start time'] = time_day[x].value[1]
            else:
                trainers[time_trainers[x].value]['t'] = 0
                trainers[time_trainers[x].value]['start time'] = time_day[x].value
        print('SCHEDGULES FROM ', day, ' HAVE BEEN ADDED')
        print('TRAINERS THIS WEEK ARE')
        for x in trainers.keys():
            if trainers[x]['t'] == 1:
                print(x)
    except TypeError:
        print(day, 'IS AN INVALID DAY, VALID DAYS ARE ', all_days)
        exit()


def pop_manuals():
    print('<ENTER THESE EMAILS IN SCHEDGULING ASSISTANT ON OUTLOOK>')
    num_manuals = 0
    for x in sp_master.keys():
        if not sp_ws.cell(row=sp_master[x]['row']+1, column =12).value or 'manual' in sp_ws.cell(row=sp_master[x]['row']+1, column =12).value:
            if not sp_ws.cell(row=sp_master[x]['row']+1, column =11).value or 'no' in sp_ws.cell(row=sp_master[x]['row']+1, column =11).value:
                if not sp_ws.cell(row=sp_master[x]['row']+1, column =9).value or 'no' in sp_ws.cell(row=sp_master[x]['row']+1, column =9).value:
                    print(sp_ws.cell(row=sp_master[x]['row']+1, column =5).value)
                    manual_users.update({sp_ws.cell(row=sp_master[x]['row']+1, column =5).value : {'ID': sp_ws.cell(row=sp_master[x]['row']+1, column =6).value, 'num meet': sp_ws.cell(row=sp_master[x]['row']+1, column =7).value, 'complete' : 'n', 'time' : ''}})
                    num_manuals+=1
    print('~~~~^^^^^ENTER THESE EMAILS IN SCHEDGULING ASSISTANT ON OUTLOOK^^^^^~~~~')
    print('TOTAL MANUAL USERS AVALIABLE', num_manuals)
    input('ENTER TO CONTINUE.....')
    print('PLEASE WAIT.....')
    
def fill_manuals(size, time, wb):
    #if size >= is_light_time(time):
        #print(time, 'has a large enough attendance with', size, 'people')
        #return
    
    print('y (add to meeting)')
    print('n (dont add)')
    print('x(to remove this user from the list)')
    print('q(to quit)')
    for x in manual_users.keys():
        
        #checks ammont of users left 
        if(size == 0):
            print(time, 'filled with max number of users')
            break
        #checks if they are done
        if 'y' in manual_users[x]['complete']:
            continue
        
        print('is ', email_to_name_LF(x), x, 'avaliable for ', time, '(', size, 'users needed still)')
        i = input()
        if 'y' in i:
            #update workbook
            wb.append([email_to_name(x),
                       x, 
                       manual_users[x]['ID'],
                       manual_users[x]['num meet'],
                       ])
            #update list of manal users 
            manual_users[x]['complete'] = 'y'
            manual_users[x]['time'] = time
            #update sharepoint with time
            update_sp_time(x, time)
            
            print(email_to_name(x), 'sucessfully added to', time, current_date)
            size -= 1
        elif 'x' in i:
            manual_users[x]['complete'] = 'x'
        elif 'q' in i:
            break

def update_sp_time(key, time):
    sp_ws.cell(row=sp_master[key]['row']+1, column=10).value = time
    sp_ws.cell(row=sp_master[key]['row']+1, column=9).value = current_date
    
def email_to_name(m):
    name = ''
    m = m.split('@')[0]
    for x in m.split('.'):
        name += (x.capitalize() +' ')
    return name
                
def email_to_name_LF(m):
    name = ''
    m = m.split('@')[0]
    for x in reversed(m.split('.')):
        name += (x.capitalize() +' ')
    return name

def check_valid_file(fnam):
    if '-' not in fnam or ' ' in fnam or '.xlsx' not in fnam or '~' in fnam:
        return False
    else:
        return True

def pack_structure():
    for x in range(len(num_meeting)):
        sp_item = {}
        sp_item.update({'num_meeting': num_meeting[x]})
        sp_item.update({'row': x})
        sp_master.update({EMAIL[x].value.lower(): sp_item})

def insert_columns(ws):
    for y in reversed(range(1, 23)):
        col = ws[ascii_uppercase[y]]
        for x in range(16, len(col)+1):
            if y < 4:
                ws.cell(row=x, column=y+3).value = ''
            else:
                ws.cell(row=x, column=y+3).value = ws.cell(row=x, column=y).value
        
    ws.cell(row=16, column=4).value = '# of Meetings'
    ws.cell(row=16, column=5).value = 'Trainer'
    ws.cell(row=16, column=6).value = 'Attendance'
    


def sort(ws):
    for x in range(16, len(ws['D'])):
        for y in range(16, len(ws['D'])):
            #print(ws.cell(row=x+1, column=4).value, ws.cell(row=y+1, column=4).value, type(ws.cell(row=y+1, column=4).value), type(ws.cell(row=y+1, column=4).value), ws.cell(row=x+1, column=1).value, ws.cell(row=y+1, column=1).value)
            try:
                if int(ws.cell(row=x+1, column=4).value) < int(ws.cell(row=y+1, column=4).value):
                    #print('swap')
                    swap_row(ws,x,y)
            except TypeError as e:
                print(e)
                
                
def swap_row(ws, r1, r2):
    for x in range(1, 25):
        temp = ws.cell(row=r1+1, column=x).value
        ws.cell(row=r1+1, column=x).value = ws.cell(row=r2+1, column=x).value
        ws.cell(row=r2+1, column=x).value = temp
        
def get_meetnum(ws):
    for x in range(16, len(ws['B'])):
        try:
            #print(sp_master[ws['B'][x].value.lower()]['num_meeting'].value, ws['A'][x].value)
            if not sp_master[ws['B'][x].value.lower()]['num_meeting'].value:
                ws.cell(row=x+1, column=4).value = '0'
            else:
                ws.cell(row=x+1, column=4).value = int(sp_master[ws['B'][x].value.lower()]['num_meeting'].value)
        except KeyError as e:
            print(e)
            ws.cell(row=x+1, column=4).value = '0'
    
def sharepoint_entry_items():
    f = open(month_folder+'/June/zak/enterInSP.txt', 'w+')
    
    for x in manual_users.keys():
        if manual_users[x]['complete'] == 'y':
            f.write('------------------------')
            f.write(x)
            f.write('\n')
            f.write(manual_users[x]['ID'])
            f.write('\n')
            f.write(current_date)
            f.write('\n')
            f.write(manual_users[x]['time'])
            f.write('\n')
    f.close()

def select_trainers():
    for x in trainers.keys():
        print('Is ', x, ' a trainer for ', current_date)
        i = input()
        if i == 'y':
            trainers[x]['t'] = 1
            print('1 for 8-5\n 2 for 9-6')
            i = input()
            if i == 1:
                trainers[x]['start time'] = '8-5'
            else:
                trainers[x]['start time'] = '9-6'
        else:
            print('Is ', x, ' a special assist for ', current_date)
            i = input()
            if i == 'y':
                trainers[x]['t'] = 2
                print('1 for 8-5\n 2 for 9-6')
                i = input()
                if i == 1:
                    trainers[x]['start time'] = '8'
                else:
                    trainers[x]['start time'] = '9'
                
#for manual trainer selection
def fill_trainers(ws, time):
    m_trains = []
    m_assists = []
    m_num = 0
    m_index = 0
    for x in trainers:
        if trainers[x]['t'] == 1 and compare_time(trainers[x]['start time'], time):
            m_trains.append(x)
        elif trainers[x]['t'] == 2 and compare_time(trainers[x]['start time'], time):
            m_assists.append(x)

    shuffle(m_trains)
    shuffle(m_assists)


    m_trains.extend(m_assists)
    print(m_trains)
    
    for x in range(16, len(ws['E'])):
        m_num += 10 + int(ws['D'][x].value)
        try:
            if m_num >= 80:
                ws.cell(row=x+1, column=5).value = m_trains[m_index]
                trainers[m_trains[m_index]]['num_meetings'] += int(ws['D'][x].value)
                trainers[m_trains[m_index]]['total_res'] += 1
                m_num = 0
                m_index += 1
            else:
                ws.cell(row=x+1, column=5).value = m_trains[m_index]
                trainers[m_trains[m_index]]['num_meetings'] += int(ws['D'][x].value)
                trainers[m_trains[m_index]]['total_res'] += 1
        except IndexError:
            print('index error')
            
def compare_time(x,y):
    if ('8-9' in y) and '9' in x:
        return False
    elif ('1-2' in y) and '9' in x:
        return False
    elif ('11-12' in y) and '8' in x:
        return False
    elif ('5-6pm' in y) and '8' in x:
        return False
    else:
        return True
    
def is_light_time(hour):
    if '8-9' in hour:
        return 10
    elif '12-1' in hour:
        return 10
    elif '1-2' in hour:
        return 10
    elif '5-6' in hour:
        return 10
    else:
        return 16

def print_trainer_data():
    for x in trainers.keys():
        print(x)
        print('Job this week and scgedgule (0 absent)(1 trainer)(2 not trainer) :', trainers[x]['t'], trainers[x]['start time'])
        print('Total number of meetings all of their attendies have :', trainers[x]['num_meetings'])
        print('Total number of attendies they are handling :', trainers[x]['total_res'])
        
pack_times(date_to_sch.split('/')[2])
#ext = input("path extension:")
ext = date_to_sch

backup_ext = '/June/zak/formateed/'
save_ext = date_to_sch
pack_structure()

pop_manuals()


for x in os.listdir(Path(month_folder+ext)):
    if check_valid_file(os.fsdecode(x)):
        print('TIME -------------- ', os.fsdecode(x), '---------------')
        sc_wb = load_workbook(month_folder+ext+'/'+x)
        sc_ws = sc_wb.get_sheet_by_name(sc_wb.get_sheet_names()[0])
        print('PACKING SCRUCTURES ...')
        insert_columns(sc_ws)
        sc_wb.save(month_folder+save_ext+os.fsdecode(x))
        fill_manuals(16, os.fsdecode(x).split('.')[0], sc_ws)
        sc_wb.save(month_folder+save_ext+os.fsdecode(x))
        print('UPDATING DOCS FOR SELECTED MANUAL USERS ...')
        get_meetnum(sc_ws)
        sc_wb.save(month_folder+save_ext+os.fsdecode(x))
        sort(sc_ws)
        sc_wb.save(month_folder+save_ext+os.fsdecode(x))
        fill_trainers(sc_ws, os.fsdecode(x.split('.')[0]))
        sc_wb.save(month_folder+save_ext+os.fsdecode(x))
        print(os.fsdecode(x), 'COMPLETED')
sp_wb.save(month_folder+ext+'updated_sp.xlsx')
sharepoint_entry_items()
print_trainer_data()