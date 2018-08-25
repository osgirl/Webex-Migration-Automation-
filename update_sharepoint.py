# -*- coding: utf-8 -*-
"""
Created on Tue May 22 11:55:06 2018

@author: zachary.shaver
"""
import os 
from os import fsdecode as dec
from pathlib import Path
import win32com.client
from openpyxl import load_workbook
from tkinter import *



email_list = {'Khendra': "khendra.davidson@baesystems.com",
              'Dylan': 'dylan.drum@baesystems.com', 
              'Dontavius': 'dontavius.hopkins@baesystems.com', 
              'Jaylan': 'jaylan.mobley@baesystems.com', 
              'Yuliya': 'yuliya.pinchuk@baesystems.com', 
              'Zachary': 'zachary.shaver@baesystems.com', 
              'Sean': 'sean.stewart2@baesystems.com', 
              'Alina': 'alina.svarishchuk@baesystems.com', 
              'George': 'george.vargas@baesystems.com', 
              'Darnell': 'darnell.wallace@baesystems.com',
              'Brandon': 'brandon.wells@baesystems.com'}



share_path = Path('//wtooleast.nosc.na.baesystems.com/Network_Documentation/Enterprise_Documentation/Voice/FedRAMP Migration 2018/Sign Up  Lists/')
sp_path = "//wtooleast.nosc.na.baesystems.com/Network_Documentation/Enterprise_Documentation/Voice/FedRAMP Migration 2018/Scripts\put_share_point_here/sharepoint.xlsx"

months = {}
day = {}


def print_structure():
    for x in months.keys():
        for y in months[x].keys():
            print(x,y,list(months[x][y]))
            
def check_valid_file(fnam):
    if '-' not in fnam or ' ' in fnam or '.xlsx' not in fnam or '~' in fnam:
        return False
    else:
        return True
def pack_dt():
#get the structure from the folder
    for month in os.listdir(share_path):
        if '.' not in dec(month):
            months.update({dec(month) : {} })
            for date in os.listdir(str(share_path)+"/"+dec(month)):
                months[dec(month)].update({dec(date): [] })
                for time in os.listdir(str(share_path)+"/"+dec(month)+"/"+dec(date)+"/"):
                    if not check_valid_file(dec(time)):
                        continue
                    months[dec(month)][dec(date)].append(dec(time))


class selection_form(Tk):
    
    def __init__(self, root):
        Tk.__init__(self, root)
        self.root = root
        self.grid()
        self.dex = 0
        self.dropVars = []
        self.dropMenus = []
        
    def itterate(self):
        self.dex += 1
        
    def make_but(self):
        self.button = Button(self.root, text='generate', command=self.complete_form).grid(column=1)
        
    def fill_lists(self, opts, t):
        self.dropVars.append(StringVar())
        self.dropVars[self.dex].set(t)
        self.dropMenus.append(OptionMenu(self, self.dropVars[self.dex], *opts, command=self.func))
        self.dropMenus[self.dex].grid(column=1,row=(self.dex*4+3))
    
    def make_check(self, atten):
        self.checks = []
        self.checkVars = []
        for x in range(len(atten)):
            self.checkVars.append(IntVar())
            self.checks.append(Checkbutton(self.root, text=atten[x], variable=self.checkVars[x]))
            self.checks[x].grid(column=1, row=(x*3))
    
    def make_but_2(self):
        self.buttom = Button(self.root, text='submit', command=self.submit_data).grid(column=1)
    
    def submit_data(self):
        self.return_data = []
        for x in range(len(self.checkVars)):
            if self.checkVars[x].get():
                self.return_data.append('Yes')
            else:
                self.return_data.append('No')
        if self.update_att():
            self.destroy()
        else:
            print("permission error someone else is using the doccument tell them to close it")
    
    def func(self, value):
        self.dropVars[self.dex]=value
        print(self.dropVars[self.dex])
        
    def complete_form(self):
        self.data = [self.dropVars[0].get(), self.dropVars[1].get(), self.dropVars[2].get(), self.dropVars[3]]
        self.destroy()
        
    def get_return_data(self):
        return self.return_data
    
    def update_att(self):
        if self.return_data:
            try:
                for x in range(len(self.return_data)):
                    timesheet.cell(row=trainer_range[x], column=6).value = self.return_data[x]
                ws.save(str(xl_path))
                return True
            except PermissionError:
                return False
    def add_user_feild(self):
        self.add_user_var = StringVar()
        self.user_feild_l = Label(text="Enter Unexpected user: (username, email, or id number)").grid(column=1)
        self.user_entry = Entry(self.root, textvariable=self.add_user_var).grid(column=1)
    
    def make_unex(self):
        self.search_unex = Button(self.root, text='add unexpected user', command=self.search_sp).grid(column=1)
         
    def search_sp(self):
        #search the sharepoint for that name 
        self.sharepoint_wb = load_workbook(sp_path)
        print('entered wb open')
        self.sharepoint_ws = self.sharepoint_wb.get_sheet_by_name(self.sharepoint_wb.get_sheet_names()[0])
        print('entered ws open')
        self.textval = self.add_user_var.get().lower()
        print('entered', self.textval)
        self.share_un = self.sharepoint_ws['B']
        self.share_e = self.sharepoint_ws['E']
        self.share_id = self.sharepoint_ws['F']
        #preform search
        print('preforming search')
        for x in range(len(self.share_un)):
            
            #if the search matches the value it will be added to the attendance sheet at the end 
            if self.textval in self.share_un[x].value.lower() or self.textval in self.share_e[x].value.lower() or self.textval in self.share_id[x].value.lower():
                # Contact Phone Number	Status	Status Date	Score	Passed	Completed	Approved by	Enrolled by	City	State	User Code	Manager	Position	Start Date
                print('user found adding their name to the attendance sheet with yes as their attendance')
                timesheet.append([self.sharepoint_ws['C'][x].value + " " + self.sharepoint_ws['D'][x].value,  
                                       self.share_e[x].value,
                                       self.share_id[x].value,
                                       self.sharepoint_ws['F'][x].value,
                                       form_data[0],
                                       'Yes',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',
                                       ' ',])
                break
            

#for creating the form to select name and stuff
def create_form(argv):
    form = selection_form(None)
    form.fill_lists(['Khendra','Dylan', 'Dontavius', 'Yuliya', 'Zachary', 'Sean', 'Alina', 'George', 'Darnell','Brandon'], 'NAME')
    form.itterate()
    form.fill_lists(['May','June', 'July'], 'MONTH')
    form.itterate()
    form.fill_lists(['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31'], 'DAY')
    form.itterate()
    form.fill_lists(['8-9am', '9-10am', '10-11am', '11-12pm', '12-1pm', '1-2pm', '2-3pm', '3-4pm', '4-5pm', '5-6pm'], 'TIME')
    form.title('ATTENDANCE SHEET')
    form.make_but()
    form.mainloop()
    return form.data    

#converting months to number from the entry
def month(m):
    if "May" in m:
        return '5_'
    elif 'June' in m:
        return '6_'
    elif 'July' in m:
        return '7_'
#pack_dt()
#print_structure()

#the create function returns the values entered and this grabs them while creating the stuff 
form_data = create_form("Enter Correct Info")


name = form_data[0]

#
xl_path = Path(str(share_path) + '/' + form_data[1] + '/' + month(form_data[1]) + form_data[2] + '/' + form_data[3] + '.xlsx')
if not xl_path.exists():
    xl_path = Path(str(share_path) + '/' + form_data[1] + '/' + month(form_data[1]) + form_data[2] + '/' + form_data[3][:-2] + 'am.xlsx')
if not xl_path.exists():
    xl_path = Path(str(share_path) + '/' + form_data[1] + '/' + month(form_data[1]) + form_data[2] + '/' + form_data[3][:-2] + '.xlsx')


print(str(xl_path), xl_path.exists())

ws = load_workbook(str(xl_path))
timesheet = ws.get_sheet_by_name(ws.get_sheet_names()[0])
trainer_names = timesheet['E']
attend = timesheet['A']
attendies = []
trainer_range = []

for x in range(16, len(trainer_names)):
    
    if trainer_names[x].value.split(' ')[0] == name:
        trainer_range.append(x+1)
        attendies.append(attend[x].value)

form = selection_form(None)
form.make_check(attendies)
form.add_user_feild()
form.make_unex()
form.make_but_2()
form.mainloop()

attendance = form.get_return_data()

#for x in range(len(attendance)):
    #timesheet.cell(row=trainer_range[x], column=6).value = attendance[x]






        



