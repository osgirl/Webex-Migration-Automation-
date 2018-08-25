# -*- coding: utf-8 -*-
"""
Created on Mon Jun 11 15:53:33 2018

@author: zachary.shaver
"""

from openpyxl import load_workbook
import os
from datetime import datetime
import re

d = datetime.now()
print(d.month, d.day)
tt = []
tt.append(d.month)
tt.append(d.day) 

total_number_complete = 0
#data structues 
trainers = {'khendra' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'dylan' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'dontavius' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'yuliya' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'zachary': {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'sean' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'alina' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'george' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'darnell' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0},
            'brandon' : {'total_meetings' : 0, 'total_migrated' : 0, 'num_meetings': 0}}

#sp_structure
sp_master = {}
#paths
sp_path = "//wtooleast.nosc.na.baesystems.com/Network_Documentation/Enterprise_Documentation/Voice/FedRAMP Migration 2018/Scripts\put_share_point_here/sharepoint.xlsx"
m_path = '//wtooleast.nosc.na.baesystems.com/Network_Documentation/Enterprise_Documentation\Voice\FedRAMP Migration 2018/Sign Up  Lists/'

#excel workbook 
sp_wb = load_workbook(sp_path)
sp_ws = sp_wb.get_sheet_by_name(sp_wb.get_sheet_names()[0])

def pack_from_time(path):
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    email = ws['B']
    print(ws['C'][15].value)
    if '# of Meetings' in ws['C'][15].value:
        nummeet = ws['C']
        trainer = ws['D']
    else:
        nummeet = ws['D']
        trainer = ws['E']
    
    date = ''
    time = ''
    for y in path.split('/'):
        if len(y) == 4 and '_' in y:
            date = y
        elif '.xlsx' in y:
            time = y.split('.')[0]
    
    for x in range(16, len(ws['A'])):
        try:
        #variables
            if not trainer[x].value:
                continue
            t_name = name_corrector(trainer[x].value.lower())
            sp_mail = email[x].value.lower()
            sp_mig =  sp_ws['K'][sp_master[sp_mail]].value
            sp_date = sp_ws['I'][sp_master[sp_mail]].value
            if not sp_mig:
                sp_mig = ''
            if not sp_date:
                sp_date = ''
            #trainer structure load 
            trainers[t_name]['total_meetings'] += 1
            try:
                if 'yes' in sp_mig or 'Yes' in sp_mig and date_match(date, sp_date):
                    trainers[t_name]['total_migrated'] += 1
                    if nummeet[x].value.isdigit():    
                        trainers[t_name]['num_meetings'] += int(nummeet[x].value)
            except AttributeError as e:
                if not nummeet[x].value:
                    continue
                trainers[t_name]['num_meetings'] += nummeet[x].value
        except KeyError as e:
            print('key err with', e)
            
        
def date_match(d1, d2):
    if not d1 or not d2:
        return False
    if str(int(d1.split('_')[1])) in d2:
        print(d1,d2)
        return True
    else:
        print(d1,d2)
        return False

def pack_spstructure(ws):
    size = len(ws['A'])
    print(size, 'users on the sharepoint this may take a while')
    for x in range(len(ws['A'])):
        sp_key = ws['E'][x].value.lower()
        if 'malito' in sp_key:
            sp_key =  sp_key.split(':')[1]
        sp_master.update({sp_key : x})
        if 'yes' in ws['K'] or 'Yes' in ws['K']:
            total_number_complete += 1
        if x % 100 == 0:
            print(x,'/',size, 'done')
        
def check_valid_file(fnam):
    if '-' not in fnam or ' ' in fnam or '.xlsx' not in fnam or '~' in fnam:
        return False
    else:
        return True

def name_corrector(name):
    if ' ' in name:
        name = name.split(' ')[0]
    if name == 'kendra' or name == 'khrenda':
        return 'khendra'
    elif name == 'yulia':
        return 'yuliya'
    elif name == 'donte' or name == 'donvatius' or name == 'dontavus':
        return 'dontavius'
    else:
        return name
    
def f_type(x):
   extension = x[-5:]
   if '.xlsx' in extension and check_valid_file(x):
       return 'x'
   elif '.' not in extension:
       return 'f'
   else:
       return 'u'

def is_past(date):
    print(date, tt[0], tt[1])
    if 'June' in date or 'May' in date or 'July' in date:
        return True
    if len(date.split('_')) < 2:
        return False
    
    month, day = date.split('_')

    if int(month) < tt[0]:
        print('lower month')
        return True
    elif int(month) == tt[0]:
        if int(day) <= tt[1]:
            print('lower month date')
            return True
    else:
        print('future date')
        return False

def fsparser(c):
    for x in os.listdir(c):
        if 'f' == f_type(os.fsdecode(x)) and is_past(os.fsdecode(x)):
            print(c+'/'+os.fsdecode(x)+'/')
            fsparser(c+'/'+os.fsdecode(x)+'/')
        elif 'x' == f_type(os.fsdecode(x)):
            pack_from_time(c+os.fsdecode(x))
        else:
            continue 
    return

def print_results():
    most_meetings = 'dontavius'
    most_mig = 'dontavius'
    most_tm = 'dontavius'
    least_meetings = 'dontavius'
    least_mig = 'dontavius'
    least_tm = 'dontavius'
    best_percent = 'dontavius'
    worst_percent = 'dontavius'
    total_script = 0
    total = 0
    for x in trainers.keys():
        mig_percents = []
        print(x)
        print('total people covered: ', trainers[x]['total_meetings'])
        print('total migrated: ', trainers[x]['total_migrated'])
        print('total number of meetings covered', trainers[x]['num_meetings'])
        print('mig percent :',trainers[x]['total_migrated']/trainers[x]['total_meetings'])
        total_script += trainers[x]['total_migrated']
        total += trainers[x]['total_meetings']
        mig_percents.append(trainers[x]['total_migrated']/trainers[x]['total_meetings'])
        if trainers[most_meetings]['total_meetings'] < trainers[x]['total_meetings']:
            most_meetings = x
        if trainers[most_mig]['total_migrated'] < trainers[x]['total_migrated']:
            most_mig = x
        if trainers[most_tm]['num_meetings'] < trainers[x]['num_meetings']:
            most_tm = x
        if trainers[least_meetings]['total_meetings'] > trainers[x]['total_meetings']:
            least_meetings = x
        if trainers[least_mig]['total_migrated'] > trainers[x]['total_migrated']:
            least_mig = x
        if trainers[least_tm]['num_meetings'] > trainers[x]['num_meetings']:
            least_tm = x
        if trainers[x]['total_migrated']/trainers[x]['total_meetings'] > trainers[best_percent]['total_migrated']/trainers[best_percent]['total_meetings']:
            best_percent = x
        if trainers[x]['total_migrated']/trainers[x]['total_meetings'] < trainers[worst_percent]['total_migrated']/trainers[worst_percent]['total_meetings']:
            worst_percent = x
    
    print('STATS:')
    print('average mig percentage: ', sum(mig_percents) / len(mig_percents))
    print('most people trained: ', most_meetings)
    print('most people migrated: ', most_mig)
    print('most trainer meetings covered: ', most_tm)
    print('least people trained: ', least_meetings)
    print('least people migrated: ', least_mig)
    print('least trainer meetings covered: ', least_tm)
    print('best migration percentage', best_percent, 'at', trainers[best_percent]['total_migrated']/trainers[best_percent]['total_meetings'])
    print('worst migration percentage', worst_percent, 'at', trainers[worst_percent]['total_migrated']/trainers[worst_percent]['total_meetings'])
    print('total number of complete : ', total_number_complete, 'number picked calculated by this script', total_script)
    print('error percentage on migration', 1 - (total_script/total_number_complete))
    print('error percentage on grabbing data', 1- (total / len(sp_ws['A'])))
    

        

pack_spstructure(sp_ws)
fsparser(m_path)
print_results()















